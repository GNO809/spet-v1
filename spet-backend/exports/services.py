# ============================================================
# SPET — exports/services.py
# Services d'export PDF & Excel des emplois du temps
# ============================================================

import io
from datetime import datetime

# ── PDF (ReportLab) ──────────────────────────────────────────
def generate_timetable_pdf(timetable) -> bytes:
    """
    Génère un PDF de l'emploi du temps au format officiel.
    Grille horaire 08H-19H, en-tête université, noms enseignants + salles.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from datetime import time as dtime
    except ImportError:
        raise ImportError("reportlab est requis pour l'export PDF. pip install reportlab")

    # ── Palette ───────────────────────────────────────────────
    C_NAVY    = colors.HexColor('#1e3a8a')
    C_BLUE    = colors.HexColor('#1a56db')
    C_RED     = colors.HexColor('#dc2626')
    C_CM      = colors.HexColor('#dbeafe')
    C_TD      = colors.HexColor('#dcfce7')
    C_TP      = colors.HexColor('#fef3c7')
    C_BREAK   = colors.HexColor('#e2e8f0')
    C_TIMECOL = colors.HexColor('#f1f5f9')
    C_GRID    = colors.HexColor('#cbd5e1')
    C_WHITE   = colors.white
    C_GREY    = colors.HexColor('#64748b')

    TYPE_FILL = {'CM': C_CM, 'TD': C_TD, 'TP': C_TP}

    # ── Créneaux fixes ────────────────────────────────────────
    SLOTS = [
        (dtime(8,  0), dtime(9,  0), '08H-09H'),
        (dtime(9,  0), dtime(10, 0), '09H-10H'),
        (dtime(10, 0), dtime(11, 0), '10H-11H'),
        (dtime(11, 0), dtime(12, 0), '11H-12H'),
        (None,          None,          '12H-15H'),   # Pause déjeuner
        (dtime(15, 0), dtime(16, 0), '15H-16H'),
        (dtime(16, 0), dtime(17, 0), '16H-17H'),
        (dtime(17, 0), dtime(18, 0), '17H-18H'),
        (dtime(18, 0), dtime(19, 0), '18H-19H'),
    ]
    LUNCH_IDX = next(i for i, (ss, _, __) in enumerate(SLOTS) if ss is None)

    DAYS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']

    # ── Séances ───────────────────────────────────────────────
    sessions = list(
        timetable.sessions
        .select_related('course', 'teacher', 'room', 'group')
        .all()
    )
    sessions_by_day = {d: [] for d in DAYS}
    for s in sessions:
        if s.day in sessions_by_day:
            sessions_by_day[s.day].append(s)

    # ── Placement dans la grille ──────────────────────────────
    # grid[slot_idx][day_idx] = ('start', sess) | ('cont', sess) | ('break',) | ('empty',)
    grid = [[('empty',) for _ in DAYS] for _ in SLOTS]
    for i, (ss, se, _) in enumerate(SLOTS):
        if ss is None:
            for j in range(len(DAYS)):
                grid[i][j] = ('break',)

    span_cmds = []

    for day_idx, day in enumerate(DAYS):
        for sess in sessions_by_day[day]:
            st, et = sess.start_time, sess.end_time

            # Trouver le premier créneau où la session débute
            first_slot = None
            for i, (ss, se, _) in enumerate(SLOTS):
                if ss is not None and ss == st:
                    first_slot = i
                    break
            if first_slot is None:
                # Fallback : premier créneau qui chevauche la session
                for i, (ss, se, _) in enumerate(SLOTS):
                    if ss is not None and ss <= st < se:
                        first_slot = i
                        break
            if first_slot is None:
                continue

            # Calculer les créneaux couverts (on ignore la pause)
            covered = []
            for i in range(first_slot, len(SLOTS)):
                ss, se, _ = SLOTS[i]
                if ss is None:
                    continue
                if ss < et:
                    covered.append(i)
                else:
                    break

            if not covered:
                covered = [first_slot]

            grid[first_slot][day_idx] = ('start', sess)
            for i in covered[1:]:
                grid[i][day_idx] = ('cont', sess)

            if len(covered) > 1:
                tr_first = first_slot + 1   # +1 pour la ligne d'en-tête
                tr_last  = covered[-1] + 1
                tc       = day_idx + 1      # +1 pour la colonne horaire
                span_cmds.append(('SPAN', (tc, tr_first), (tc, tr_last)))

    # ── Styles de paragraphe ──────────────────────────────────
    styles = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles['Normal'], **kw)

    hdr_p  = ps('hdr',  fontSize=7,  alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=0, leading=9)
    sub_p  = ps('sub',  fontSize=6,  alignment=TA_CENTER, textColor=C_GREY,          spaceAfter=0, spaceBefore=0, leading=8)
    time_p = ps('time', fontSize=7,  alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=0, leading=9)

    def cell_para(sess):
        """Paragraphe riche pour une case de la grille."""
        if sess is None:
            return ''
        course_name = sess.course.name if sess.course else '—'
        teacher_str = ''
        if sess.teacher:
            fn = (sess.teacher.first_name or '').strip()
            ln = (sess.teacher.last_name  or '').strip()
            teacher_str = f'({fn} {ln})'.strip('( )').strip()
            teacher_str = f'({teacher_str})'
        room_str = f'<font color="#dc2626">{sess.room.name}</font>' if sess.room else ''
        parts = [f'<b>{course_name}</b>']
        if teacher_str:
            parts.append(f'<font size="5.5">{teacher_str}</font>')
        if room_str:
            parts.append(f'<font size="5.5">{room_str}</font>')
        cell_style = ps(
            f'cell_{id(sess)}',
            fontSize=6.5, alignment=TA_CENTER,
            leading=8.5, spaceBefore=0, spaceAfter=0,
            allowWidows=0, allowOrphans=0,
        )
        return Paragraph('<br/>'.join(parts), cell_style)

    # ── Construction de la table ──────────────────────────────
    header_row = [Paragraph('HEURES', hdr_p)] + [Paragraph(d, hdr_p) for d in DAYS]
    data = [header_row]

    ROW_H = []
    ROW_H.append(0.65 * cm)   # header

    pause_style = ps('pause', fontSize=7, fontName='Helvetica-Oblique',
                     alignment=TA_CENTER, textColor=C_GREY,
                     spaceBefore=0, spaceAfter=0, leading=9)

    for slot_idx, (ss, se, label) in enumerate(SLOTS):
        row = [Paragraph(label, time_p)]
        for day_idx in range(len(DAYS)):
            cell = grid[slot_idx][day_idx]
            if cell[0] == 'start':
                row.append(cell_para(cell[1]))
            elif cell[0] == 'break' and day_idx == 0:
                row.append(Paragraph('— Pause déjeuner —', pause_style))
            elif cell[0] in ('cont', 'break', 'empty'):
                row.append('')
        data.append(row)
        ROW_H.append(0.45 * cm if ss is None else 1.55 * cm)

    # ── Largeurs des colonnes ─────────────────────────────────
    PAGE_W = landscape(A4)[0] - 2.4 * cm
    TIME_W = 1.7 * cm
    DAY_W  = (PAGE_W - TIME_W) / 6
    col_widths = [TIME_W] + [DAY_W] * 6

    table = Table(data, colWidths=col_widths, rowHeights=ROW_H, repeatRows=1)

    # ── TableStyle ────────────────────────────────────────────
    ts = [
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), C_NAVY),
        ('TEXTCOLOR',  (0, 0), (-1, 0), C_WHITE),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, 0), 'MIDDLE'),
        # Colonne horaire
        ('BACKGROUND', (0, 1), (0, -1), C_TIMECOL),
        ('FONTNAME',   (0, 1), (0, -1), 'Helvetica-Bold'),
        ('ALIGN',      (0, 1), (0, -1), 'CENTER'),
        ('VALIGN',     (0, 1), (0, -1), 'MIDDLE'),
        # Cases normales
        ('ALIGN',      (1, 1), (-1, -1), 'CENTER'),
        ('VALIGN',     (1, 1), (-1, -1), 'MIDDLE'),
        # Grille
        ('GRID',       (0, 0), (-1, -1), 0.5, C_GRID),
        # Padding
        ('TOPPADDING',    (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING',   (0, 0), (-1, -1), 3),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 3),
    ]

    # Pause déjeuner
    lunch_row = LUNCH_IDX + 1
    ts += [
        ('BACKGROUND', (0, lunch_row), (-1, lunch_row), C_BREAK),
        ('FONTNAME',   (0, lunch_row), (-1, lunch_row), 'Helvetica-Oblique'),
        ('SPAN',       (1, lunch_row), (-1, lunch_row)),
    ]

    # Couleur par type de séance + fond blanc pour les cases vides
    for slot_idx in range(len(SLOTS)):
        for day_idx in range(len(DAYS)):
            cell = grid[slot_idx][day_idx]
            tr = slot_idx + 1
            tc = day_idx + 1
            if cell[0] == 'start':
                fill = TYPE_FILL.get(cell[1].session_type, C_CM)
                ts.append(('BACKGROUND', (tc, tr), (tc, tr), fill))
            elif cell[0] == 'empty':
                ts.append(('BACKGROUND', (tc, tr), (tc, tr), C_WHITE))

    # SPAN multi-créneaux
    ts.extend(span_cmds)

    table.setStyle(TableStyle(ts))

    # ── En-tête du document ───────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.0 * cm,
    )

    year_label = (
        timetable.academic_year.label
        if hasattr(timetable, 'academic_year') and timetable.academic_year
        else str(datetime.now().year)
    )

    def hp(text, **kw):
        style_kw = {'alignment': TA_CENTER, 'spaceBefore': 0, 'spaceAfter': 2}
        style_kw.update(kw)
        return Paragraph(text, ps(f'hp_{id(text)}', **style_kw))

    elements = [
        hp('REPUBLIQUE DU SENEGAL',  fontSize=8,  fontName='Helvetica-Bold'),
        hp('Un Peuple - Un But - Une Foi', fontSize=7),
        hp("Ministère de l'Enseignement supérieur, de la Recherche et de l'Innovation", fontSize=6.5),
        hp('UNIVERSITE IBA DER THIAM DE THIES', fontSize=9, fontName='Helvetica-Bold', spaceAfter=4),
        HRFlowable(width='100%', thickness=1, color=C_NAVY, spaceAfter=4),
        hp(f'ANNEE {year_label} — {timetable.semestre}',
           fontSize=11, fontName='Helvetica-Bold', spaceAfter=2),
        hp(f'EMPLOI DU TEMPS — {timetable.filiere.name.upper()}',
           fontSize=13, fontName='Helvetica-Bold', textColor=C_RED, spaceAfter=2),
        hp(f'Généré le {datetime.now().strftime("%d/%m/%Y à %H:%M")}',
           fontSize=6.5, textColor=C_GREY, spaceAfter=6),
        table,
        Spacer(1, 0.3 * cm),
        hp('CM : Cours Magistral  |  TD : Travaux Dirigés  |  TP : Travaux Pratiques',
           fontSize=6, textColor=C_GREY),
    ]

    doc.build(elements)
    return buffer.getvalue()


# ── Excel (openpyxl) ─────────────────────────────────────────
def generate_timetable_excel(timetable) -> bytes:
    """
    Génère un fichier Excel de l'emploi du temps.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, Fill, PatternFill, Alignment,
            Border, Side, GradientFill,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl est requis. pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = f'EDT {timetable.filiere.code}'

    DAYS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    sessions = timetable.sessions.select_related('course', 'teacher', 'room', 'group').all()

    sessions_by_day = {d: [] for d in DAYS}
    for s in sessions:
        if s.day in sessions_by_day:
            sessions_by_day[s.day].append(s)

    time_pairs = sorted(set((s.start_time, s.end_time) for s in sessions))

    # ── Couleurs ─────────────────────────────────────────────
    HDR_FILL = PatternFill('solid', fgColor='1e3a8a')
    HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
    CM_FILL  = PatternFill('solid', fgColor='dbeafe')
    TD_FILL  = PatternFill('solid', fgColor='dcfce7')
    TP_FILL  = PatternFill('solid', fgColor='fef3c7')
    TYPE_FILL = {'CM': CM_FILL, 'TD': TD_FILL, 'TP': TP_FILL}

    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0'),
    )

    # ── Titre ────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell = ws.cell(row=1, column=1,
        value=f'Emploi du Temps — {timetable.filiere.name} — {timetable.semestre}')
    title_cell.font      = Font(bold=True, size=13, color='1e3a8a')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── En-têtes ─────────────────────────────────────────────
    headers = ['Créneau'] + DAYS
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = thin_border

    ws.column_dimensions['A'].width = 14
    for col_idx in range(2, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.row_dimensions[2].height = 24

    # ── Données ──────────────────────────────────────────────
    for row_idx, (start, end) in enumerate(time_pairs, start=3):
        ws.row_dimensions[row_idx].height = 55
        time_cell = ws.cell(
            row=row_idx, column=1,
            value=f'{start.strftime("%H:%M")} – {end.strftime("%H:%M")}',
        )
        time_cell.font      = Font(bold=True, size=9, color='1e293b')
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        time_cell.border    = thin_border

        for col_idx, day in enumerate(DAYS, start=2):
            day_sessions = [
                s for s in sessions_by_day[day]
                if s.start_time == start and s.end_time == end
            ]
            if day_sessions:
                s         = day_sessions[0]
                cell_val  = (
                    f'{s.course.name}\n'
                    f'[{s.session_type}]\n'
                    f'{s.teacher.full_name if s.teacher else "—"}\n'
                    f'{s.room.name if s.room else "—"}'
                )
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                cell.fill = TYPE_FILL.get(s.session_type, CM_FILL)
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value='')

            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font      = Font(size=8)
            cell.border    = thin_border

    # ── Feuille de statistiques ───────────────────────────────
    ws2 = wb.create_sheet('Statistiques')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    ws2.cell(row=1, column=1, value='Filière').font = Font(bold=True)
    ws2.cell(row=1, column=2, value=timetable.filiere.name)
    ws2.cell(row=2, column=1, value='Semestre').font = Font(bold=True)
    ws2.cell(row=2, column=2, value=timetable.semestre)
    ws2.cell(row=3, column=1, value='Nombre de séances').font = Font(bold=True)
    ws2.cell(row=3, column=2, value=timetable.sessions.count())
    ws2.cell(row=4, column=1, value='Score qualité').font = Font(bold=True)
    ws2.cell(row=4, column=2, value=f'{timetable.quality_score}%')
    ws2.cell(row=5, column=1, value='Statut').font = Font(bold=True)
    ws2.cell(row=5, column=2, value=timetable.get_status_display())

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Export enseignant ─────────────────────────────────────────
def generate_teacher_timetable_pdf(teacher, sessions) -> bytes:
    """PDF personnel d'un enseignant (ses séances uniquement)."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        raise ImportError("reportlab est requis.")

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Mon Emploi du Temps — {teacher.full_name}", styles['Title']),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']),
        Spacer(1, 0.5*cm),
    ]

    data = [['Jour', 'Début', 'Fin', 'Cours', 'Type', 'Salle', 'Groupe']]
    for s in sessions:
        data.append([
            s.day,
            s.start_time.strftime('%H:%M'),
            s.end_time.strftime('%H:%M'),
            s.course.name,
            s.session_type,
            s.room.name if s.room else '—',
            s.group.label if s.group else '—',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',    (0,0), (-1,-1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()
